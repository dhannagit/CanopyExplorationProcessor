"""Generate turn-zone masks from the Canopy circuit workbook."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

from .dxpx import DXPXRun


class TurnZoneError(RuntimeError):
    """Raised when turn zones cannot be generated for a run."""


def add_turn_zones(run: DXPXRun, workbook_path: str | Path) -> DXPXRun:
    """Return a run with missing turn-zone masks populated.

    Existing ``isTurnZoneN`` fields are retained exactly. If no masks exist,
    the circuit workbook is used with the same inclusive ``sLap`` bounds as
    the MATLAB ``create_TurnZone`` implementation.
    """

    existing = run.turn_zone_names
    if existing:
        return run

    lap_distance = _as_vector(run.data.get("sLap"), "sLap")
    track_name = run.track_name
    if not track_name:
        raise TurnZoneError(f"Run has no Header.TrackName: {run.path}")

    zones = _read_circuit_zones(workbook_path, track_name)
    updated_data = deepcopy(run.data)
    for index, (start, end) in enumerate(zones, start=1):
        updated_data[f"isTurnZone{index}"] = (
            (lap_distance >= start) & (lap_distance <= end)
        ).astype(float)

    return DXPXRun(
        path=run.path,
        header=run.header,
        data=updated_data,
        unit=run.unit,
        type_info=run.type_info,
    )


def _read_circuit_zones(workbook_path: str | Path, track_name: str) -> list[tuple[float, float]]:
    workbook = Path(workbook_path)
    if not workbook.is_file():
        raise TurnZoneError(f"Circuit workbook does not exist: {workbook}")

    sheet = load_workbook(workbook, read_only=True, data_only=True)["Turns"]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value) if value is not None else "" for value in next(rows)]
    except StopIteration as exc:
        raise TurnZoneError(f"Circuit workbook has no header row: {workbook}") from exc

    circuit_row: tuple[Any, ...] | None = None
    for row in rows:
        if row and str(row[0]).strip() == track_name.strip():
            circuit_row = row
            break
    if circuit_row is None:
        raise TurnZoneError(f"Circuit {track_name!r} is not present in {workbook}")

    zones: list[tuple[float, float]] = []
    for index in range(1, len(headers), 2):
        if index + 1 >= len(headers):
            break
        start = circuit_row[index] if index < len(circuit_row) else None
        end = circuit_row[index + 1] if index + 1 < len(circuit_row) else None
        if start is None and end is None:
            continue
        if start is None or end is None:
            raise TurnZoneError(f"Incomplete turn zone {headers[index]} for circuit {track_name!r}")
        zones.append((float(start), float(end)))

    if not zones:
        raise TurnZoneError(f"Circuit {track_name!r} has no turn zones in {workbook}")
    return zones


def _as_vector(value: Any, field_name: str) -> np.ndarray:
    if value is None:
        raise TurnZoneError(f"Run has no Data.{field_name} field")
    vector = np.asarray(value).reshape(-1)
    if vector.size == 0 or not np.issubdtype(vector.dtype, np.number):
        raise TurnZoneError(f"Data.{field_name} must be a non-empty numeric vector")
    return vector